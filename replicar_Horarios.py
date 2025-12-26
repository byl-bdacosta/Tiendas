#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import sys
from pathlib import Path
from typing import Dict, Tuple, List

import openpyxl
from lxml import etree as ET


DW_NS = "http://www.demandware.com/xml/impex/store/2007-04-30"
XML_NS = "http://www.w3.org/XML/1998/namespace"
NSMAP = {"dw": DW_NS, "xml": XML_NS}


def eprint(msg: str) -> None:
    print(msg, file=sys.stderr)


def read_excel_hours(
    xlsx_path: Path,
    sheet_name: str,
    header_row: int = 3,
    data_start_row: int = 4,
    debug: bool = False,
    debug_n: int = 3,
) -> Dict[str, Dict[str, str]]:
    """
    Lee la hoja indicada y devuelve:
      { "C03": {"x-default": "...", "ca": "...", "gl": "...", "es": "..."}, ... }

    - Detecta las columnas por el texto de la fila de cabecera (header_row).
    - Conserva saltos de línea (openpyxl los devuelve en el string).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe. Hojas: {wb.sheetnames}")

    ws = wb[sheet_name]

    headers: Dict[str, int] = {}
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        v = ws.cell(header_row, col).value
        if v is None:
            continue
        key = str(v).strip().lower()
        if key:
            headers[key] = col

    def find_col(candidates: list[str]) -> int:
 
        for c in candidates:
            c_norm = c.strip().lower()
            if c_norm in headers:
                return headers[c_norm]

        for h, col in headers.items():
            for c in candidates:
                if c.strip().lower() in h:
                    return col
        raise ValueError(
            f"No encuentro columna para: {candidates}.\n"
            f"Headers detectados (primeros 40): {list(headers.keys())[:40]}"
        )

    col_cod = find_col(["cod", "cód", "código","COD"])
    col_en = find_col(["inglés", "ingles", "default", "x-default"])
    col_ca = find_col(["catalán", "catala", "català", "ca"])
    col_gl = find_col(["gallego", "galego", "gl", "ga"])
    col_es = find_col(["español", "espanol", "castellano", "es"])

    out: Dict[str, Dict[str, str]] = {}
    row = data_start_row
    debug_count = 0

    def norm(v) -> str:
        if v is None:
            return ""
        return str(v)

    while True:
        cod = ws.cell(row, col_cod).value
        if cod is None or str(cod).strip() == "":
            break

        store_id = str(cod).strip()

        payload = {
            "x-default": norm(ws.cell(row, col_en).value),
            "ca": norm(ws.cell(row, col_ca).value),
            "gl": norm(ws.cell(row, col_gl).value),
            "es": norm(ws.cell(row, col_es).value),
        }

        if debug and debug_count < debug_n:
            print(
                f"[DEBUG] {store_id} -> "
                f"x-default({len(payload['x-default'])} chars), "
                f"ca({len(payload['ca'])}), "
                f"gl({len(payload['gl'])}), "
                f"es({len(payload['es'])})"
            )
            debug_count += 1

        if store_id in out:
            raise ValueError(f"COD duplicado en Excel: {store_id} (fila {row})")

        out[store_id] = payload
        row += 1

    return out


def index_stores(root: ET._Element) -> Dict[str, ET._Element]:
    stores: Dict[str, ET._Element] = {}
    for store in root.xpath("//dw:store", namespaces=NSMAP):
        sid = store.get("store-id")
        if sid:
            stores[sid] = store
    return stores


def find_store_hours_node(store_el: ET._Element, lang: str) -> ET._Element | None:
    found = store_el.xpath(f'./dw:store-hours[@xml:lang="{lang}"]', namespaces=NSMAP)
    return found[0] if found else None


def update_xml(
    xml_path: Path,
    hours_by_store: Dict[str, Dict[str, str]],
    strict: bool = False,
) -> Tuple[ET._ElementTree, Dict[str, int], List[str]]:
    parser = ET.XMLParser(remove_blank_text=False, recover=False, huge_tree=True)
    tree = ET.parse(str(xml_path), parser)
    root = tree.getroot()

    store_index = index_stores(root)

    summary = {
        "stores_in_excel": len(hours_by_store),
        "stores_found_in_xml": 0,
        "stores_missing_in_xml": 0,
        "store_hours_updated": 0,
        "store_hours_missing_nodes": 0,
    }
    errors: List[str] = []

    langs_to_update = ["x-default", "ca", "gl", "es"]

    for store_id, lang_texts in hours_by_store.items():
        store_el = store_index.get(store_id)
        if store_el is None:
            msg = f"ERROR: store-id '{store_id}' está en Excel pero no existe en el XML."
            errors.append(msg)
            eprint(msg)
            summary["stores_missing_in_xml"] += 1
            continue

        summary["stores_found_in_xml"] += 1

        for lang in langs_to_update:
            node = find_store_hours_node(store_el, lang)
            if node is None:
                msg = f"ERROR: store-id '{store_id}' no tiene <store-hours xml:lang=\"{lang}\"> en el XML."
                errors.append(msg)
                eprint(msg)
                summary["store_hours_missing_nodes"] += 1
                continue

            node.text = lang_texts.get(lang, "")
            summary["store_hours_updated"] += 1

        ga_node = find_store_hours_node(store_el, "ga")
        if ga_node is not None:
            ga_node.text = lang_texts.get("gl", "")
            summary["store_hours_updated"] += 1

    if strict and errors:
        raise RuntimeError(f"Se encontraron errores ({len(errors)}).")

    return tree, summary, errors


def write_xml(tree: ET._ElementTree, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tree.write(
        str(out_path),
        encoding="UTF-8",
        xml_declaration=True,
        pretty_print=False,  
    )


def main() -> int:
    p = argparse.ArgumentParser(
        description="Actualiza store-hours en stores.xml usando horarios de un Excel (HORARIO INTERNACIONAL)."
    )
    p.add_argument("--xml", required=True, type=Path, help="Ruta al XML de tiendas (stores.xml)")
    p.add_argument("--xlsx", required=True, type=Path, help="Ruta al Excel de horarios")
    p.add_argument("--sheet", default="HORARIO INTERNACIONAL", help="Nombre de la hoja (default: HORARIO INTERNACIONAL)")

    out_group = p.add_mutually_exclusive_group()
    out_group.add_argument("--inplace", action="store_true", help="Sobrescribe el XML de entrada")
    out_group.add_argument("--out", type=Path, help="Escribe el XML actualizado en otra ruta")

    p.add_argument("--dry-run", action="store_true", help="No escribe archivo, solo muestra resumen")
    p.add_argument("--strict", action="store_true", help="Falla (exit!=0) si hay cualquier ERROR")
    p.add_argument("--debug", action="store_true", help="Imprime qué está leyendo del Excel (primeras tiendas)")
    p.add_argument("--debug-n", type=int, default=3, help="Número de tiendas a imprimir con --debug")

    args = p.parse_args()

    if not args.xml.exists():
        eprint(f"ERROR: no existe el XML: {args.xml}")
        return 2
    if not args.xlsx.exists():
        eprint(f"ERROR: no existe el XLSX: {args.xlsx}")
        return 2

    hours_by_store = read_excel_hours(
        args.xlsx,
        args.sheet,
        debug=args.debug,
        debug_n=args.debug_n,
    )

    try:
        tree, summary, errors = update_xml(args.xml, hours_by_store, strict=args.strict)
    except Exception as ex:
        eprint(f"ERROR: {ex}")
        return 3

    print("=== RESUMEN ===")
    for k, v in summary.items():
        print(f"{k}: {v}")
    print(f"errors: {len(errors)}")

    if args.dry_run:
        print("dry-run: no se escribió ningún archivo.")
        return 0 if (not args.strict or len(errors) == 0) else 4

    if args.inplace:
        out_path = args.xml
    else:
        out_path = args.out if args.out else args.xml.with_suffix(".updated.xml")

    write_xml(tree, out_path)
    print(f"OK: XML escrito en: {out_path}")
    return 0 if (not args.strict or len(errors) == 0) else 4


if __name__ == "__main__":
    raise SystemExit(main())
